from datetime import datetime
from io import BytesIO
from copy import copy
from pathlib import Path
import re
import shutil

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

from app.config import ARCHIVE_DIR
from app.importers.topos_importer import caminho_sites_excel


SITE_REGISTRY_COLUMNS = [
    "CÓDIGO AQUILES",
    "CÓDIGO MICROSIGA",
    "CÓDIGO CONDOMINIO",
    "ABREVIAÇÃO",
    "SMNPC",
    "TIPO",
    "NOME",
    "Relacionamento",
    "Favorecido",
    "CONTRATO",
    "QTDO",
    "CATEGORIA",
    "PERFIL",
    "CUSTO",
    "ENDEREÇO",
    "NUMERO",
    "BAIRRO",
    "CIDADE",
    "UF",
    "CEP",
    "ATIVAÇÃO",
    "LATITUDE",
    "LONGITUDE",
    "ALTURA",
    "RESTRIÇÃO",
    "Status",
    "Detalhe",
    "OBSERVAÇÃO:"
]

SITE_CONTACT_COLUMNS = [
    "CÓDIGO AQUILES",
    "Tipo de contato",
    "Nome",
    "Telefones",
    "Emails",
    "Observações",
    "Arquivado",
    "Arquivado em",
    "Arquivado por"
]

SITE_CONTACT_TYPES = [
    "Sindico",
    "Zelador",
    "Administradora",
    "Portaria",
    "Morador",
    "Proprietário",
    "Outro"
]

NUMERIC_COLUMNS = {
    "QTDO",
    "ALTURA"
}
CURRENCY_COLUMNS = {
    "CUSTO"
}

SITE_CODE_COLUMN = "CÓDIGO AQUILES"
SITE_CONTACTS_SHEET = "CONTATOS"

UNIQUE_SITE_CODE_COLUMNS = [
    "CÓDIGO AQUILES",
    "CÓDIGO MICROSIGA",
    "CÓDIGO CONDOMINIO",
    "ABREVIAÇÃO",
    "SMNPC",
    "NOME",
    "Favorecido"
]

UNIQUE_SITE_CODE_LABELS = {
    "CÓDIGO AQUILES": "Código Aquiles",
    "CÓDIGO MICROSIGA": "Código Microsiga",
    "CÓDIGO CONDOMINIO": "Código Condomínio",
    "ABREVIAÇÃO": "Abreviação",
    "SMNPC": "SNMPc",
    "NOME": "Nome",
    "Favorecido": "Favorecido"
}

COLUMN_ALIASES = {
    "CÓDIGO": "CÓDIGO AQUILES",
    "CODIGO": "CÓDIGO AQUILES",
    "CODIGO AQUILES": "CÓDIGO AQUILES",
    "Microsiga": "CÓDIGO MICROSIGA",
    "MICROSIGA": "CÓDIGO MICROSIGA",
    "CODIGO MICROSIGA": "CÓDIGO MICROSIGA",
    "CODIGO CONDOMINIO": "CÓDIGO CONDOMINIO",
    "CONDOMINIO": "CÓDIGO CONDOMINIO",
    "ABREVIACAO": "ABREVIAÇÃO",
    "RELACIONAMENTO": "Relacionamento",
    "FAVORECIDO": "Favorecido"
}

CONTACT_COLUMN_ALIASES = {
    "CÓDIGO": "CÓDIGO AQUILES",
    "CODIGO": "CÓDIGO AQUILES",
    "CODIGO AQUILES": "CÓDIGO AQUILES",
    "TIPO": "Tipo de contato",
    "TIPO CONTATO": "Tipo de contato",
    "TIPO DE CONTATO": "Tipo de contato",
    "CONTATO": "Tipo de contato",
    "NOME": "Nome",
    "NOME CONTATO": "Nome",
    "NOME DO CONTATO": "Nome",
    "TELEFONE": "Telefones",
    "TELEFONES": "Telefones",
    "FONE": "Telefones",
    "FONES": "Telefones",
    "EMAIL": "Emails",
    "EMAILS": "Emails",
    "E-MAIL": "Emails",
    "E-MAILS": "Emails",
    "OBS": "Observações",
    "OBSERVACAO": "Observações",
    "OBSERVAÇÃO": "Observações",
    "OBSERVACOES": "Observações",
    "OBSERVAÇÕES": "Observações",
    "OBSERVATION": "Observações",
    "NOTAS": "Observações",
    "ARQUIVADO": "Arquivado",
    "ARQUIVADA": "Arquivado",
    "ARQUIVADO EM": "Arquivado em",
    "ARQUIVADA EM": "Arquivado em",
    "DATA ARQUIVAMENTO": "Arquivado em",
    "ARQUIVADO POR": "Arquivado por",
    "ARQUIVADA POR": "Arquivado por"
}


def normalize_code(value):
    if pd.isna(value):
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()

    if text.endswith(".0"):
        text = text[:-2]

    return text


def normalize_column_key(value):
    text = str(value or "").strip()
    text = (
        text
        .replace("Á", "A")
        .replace("À", "A")
        .replace("Ã", "A")
        .replace("Â", "A")
        .replace("É", "E")
        .replace("Ê", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Õ", "O")
        .replace("Ô", "O")
        .replace("Ú", "U")
        .replace("Ç", "C")
        .replace("á", "a")
        .replace("à", "a")
        .replace("ã", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("õ", "o")
        .replace("ô", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )

    return re.sub(r"[^A-Za-z0-9]+", " ", text).strip().upper()


def normalize_columns(df, aliases):
    rename = {}

    for column in df.columns:
        key = normalize_column_key(column)

        if key in aliases:
            rename[column] = aliases[key]

    return df.rename(columns=rename)


def canonical_column_name(column, aliases):
    key = normalize_column_key(column)

    if key in aliases:
        return aliases[key]

    for canonical in SITE_REGISTRY_COLUMNS + SITE_CONTACT_COLUMNS:
        if normalize_column_key(canonical) == key:
            return canonical

    return str(column or "").strip()


def parse_currency_value(value):
    if pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value or "").strip()

    if not text:
        return None

    text = (
        text
        .replace("R$", "")
        .replace(" ", "")
        .strip()
    )

    if "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return value


def value_for_excel(column, value):
    if pd.isna(value):
        return None

    if column in CURRENCY_COLUMNS:
        return parse_currency_value(value)

    if column in NUMERIC_COLUMNS:
        if value in ("", None):
            return None

        try:
            return float(value)
        except (TypeError, ValueError):
            return value

    return value


def copy_cell_format(source, target):
    if not source.has_style:
        return

    target._style = copy(source._style)

    if source.number_format:
        target.number_format = source.number_format

    if source.font:
        target.font = copy(source.font)

    if source.fill:
        target.fill = copy(source.fill)

    if source.border:
        target.border = copy(source.border)

    if source.alignment:
        target.alignment = copy(source.alignment)

    if source.protection:
        target.protection = copy(source.protection)


def worksheet_headers(ws):
    headers = {}

    for column_index in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=column_index).value

        if header is None:
            continue

        headers[canonical_column_name(header, COLUMN_ALIASES)] = column_index

    return headers


def ensure_worksheet_columns(ws, columns, aliases):
    headers = {}

    for column_index in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=column_index).value

        if header is None:
            continue

        headers[canonical_column_name(header, aliases)] = column_index

    for column in columns:
        if column in headers:
            continue

        column_index = (
            1
            if (
                not headers
                and ws.max_column == 1
                and ws.cell(row=1, column=1).value is None
            )
            else ws.max_column + 1
        )
        ws.cell(row=1, column=column_index).value = column

        if column_index > 1:
            copy_cell_format(
                ws.cell(row=1, column=column_index - 1),
                ws.cell(row=1, column=column_index)
            )

        headers[column] = column_index

    return headers


def registry_sheet_name(workbook):
    if "BASE" in workbook.sheetnames:
        return "BASE"

    if workbook.sheetnames:
        return workbook.sheetnames[0]

    return "BASE"


def load_or_create_workbook(path):
    if path.exists():
        return load_workbook(path)

    workbook = Workbook()
    workbook.active.title = "BASE"

    return workbook


def write_dataframe_preserving_sheet(ws, df, columns, aliases):
    headers = ensure_worksheet_columns(
        ws,
        columns,
        aliases
    )

    previous_max_row = ws.max_row

    for row_offset, (_index, row) in enumerate(df.iterrows(), start=2):
        template_row = min(
            row_offset,
            previous_max_row
        )

        if template_row < 2 and previous_max_row >= 2:
            template_row = 2

        for column in columns:
            column_index = headers[column]
            target = ws.cell(row=row_offset, column=column_index)

            if row_offset > previous_max_row and template_row >= 2:
                copy_cell_format(
                    ws.cell(row=template_row, column=column_index),
                    target
                )

            target.value = value_for_excel(
                column,
                row.get(column, "")
            )

    first_empty_row = len(df) + 2

    for row_index in range(first_empty_row, previous_max_row + 1):
        for column in columns:
            ws.cell(
                row=row_index,
                column=headers[column]
            ).value = None


def split_address_number(address):
    text = str(address or "").strip()

    if not text:
        return "", ""

    match = re.match(
        r"^(?P<address>.+?),\s*(?P<number>[^,]+)$",
        text
    )

    if not match:
        return text, ""

    return (
        match.group("address").strip(),
        match.group("number").strip()
    )


def migrate_address_number(df):
    if "ENDEREÇO" not in df.columns:
        return df

    if "NUMERO" not in df.columns:
        df["NUMERO"] = ""

    for index, row in df.iterrows():
        number = str(row.get("NUMERO") or "").strip()

        if number:
            continue

        address, parsed_number = split_address_number(
            row.get("ENDEREÇO")
        )

        if not parsed_number:
            continue

        df.at[index, "ENDEREÇO"] = address
        df.at[index, "NUMERO"] = parsed_number

    return df


def load_site_registry(path=None):
    path = Path(path) if path else caminho_sites_excel()

    if not path.exists():
        return pd.DataFrame(columns=SITE_REGISTRY_COLUMNS)

    xl = pd.ExcelFile(path)
    sheet_name = "BASE" if "BASE" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet_name, dtype=object)

    unnamed_columns = [
        column
        for column in df.columns
        if str(column).startswith("Unnamed:")
        and df[column].isna().all()
    ]

    if unnamed_columns:
        df = df.drop(columns=unnamed_columns)

    df = normalize_columns(
        df,
        COLUMN_ALIASES
    )
    df = migrate_address_number(df)

    for column in SITE_REGISTRY_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df = df[SITE_REGISTRY_COLUMNS].fillna("")
    df["CUSTO"] = df["CUSTO"].astype(str).str.strip()

    return df


def legacy_contacts_from_sites(df):
    contacts = []

    for _, row in df.iterrows():
        code = normalize_code(
            row.get(SITE_CODE_COLUMN)
        )
        contact = str(row.get("Contato Principal") or "").strip()
        phone = str(row.get("Telefone Contato") or "").strip()
        email = str(row.get("Email Contato") or "").strip()
        others = str(row.get("Outros Contatos") or "").strip()

        if contact or phone or email:
            contacts.append({
                "CÓDIGO AQUILES": code,
                "Tipo de contato": "Principal",
                "Nome": contact,
                "Telefones": phone,
                "Emails": email,
                "Observações": "",
                "Arquivado": "",
                "Arquivado em": "",
                "Arquivado por": ""
            })

        if others:
            contacts.append({
                "CÓDIGO AQUILES": code,
                "Tipo de contato": "Outros",
                "Nome": "",
                "Telefones": "",
                "Emails": others,
                "Observações": "",
                "Arquivado": "",
                "Arquivado em": "",
                "Arquivado por": ""
            })

    return pd.DataFrame(
        contacts,
        columns=SITE_CONTACT_COLUMNS
    )


def normalize_site_contacts(df):
    df = normalize_columns(
        df.copy(),
        CONTACT_COLUMN_ALIASES
    )

    for column in SITE_CONTACT_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df = df[SITE_CONTACT_COLUMNS].fillna("")
    df["CÓDIGO AQUILES"] = df["CÓDIGO AQUILES"].apply(normalize_code)

    has_site = df["CÓDIGO AQUILES"].astype(str).str.strip().ne("")
    has_contact = (
        df["Tipo de contato"].astype(str).str.strip().ne("")
        | df["Nome"].astype(str).str.strip().ne("")
        | df["Telefones"].astype(str).str.strip().ne("")
        | df["Emails"].astype(str).str.strip().ne("")
        | df["Observações"].astype(str).str.strip().ne("")
    )

    return df[has_site & has_contact].reset_index(drop=True)


def load_site_contacts(path=None):
    path = Path(path) if path else caminho_sites_excel()

    if not path.exists():
        return pd.DataFrame(columns=SITE_CONTACT_COLUMNS)

    xl = pd.ExcelFile(path)

    if SITE_CONTACTS_SHEET in xl.sheet_names:
        df = pd.read_excel(
            path,
            sheet_name=SITE_CONTACTS_SHEET,
            dtype=object
        )

        return normalize_site_contacts(df)

    sheet_name = "BASE" if "BASE" in xl.sheet_names else xl.sheet_names[0]
    df_sites = pd.read_excel(
        path,
        sheet_name=sheet_name,
        dtype=object
    ).fillna("")
    df_sites = normalize_columns(
        df_sites,
        COLUMN_ALIASES
    )

    return legacy_contacts_from_sites(df_sites)


def backup_site_registry(path=None):
    path = Path(path) if path else caminho_sites_excel()

    if not path.exists():
        return None

    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ARCHIVE_DIR / f"{path.stem}_cadastro_{suffix}{path.suffix}"
    shutil.copy2(path, backup)

    return backup


def prepare_registry_for_save(df):
    df = df.copy()

    df = migrate_address_number(df)

    for column in SITE_REGISTRY_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    df = df[SITE_REGISTRY_COLUMNS].fillna("")
    df["CUSTO"] = df["CUSTO"].astype(str).str.strip()

    for column in NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(
            df[column],
            errors="coerce"
        ).fillna(0)

    return df


def prepare_contacts_for_save(df):
    return normalize_site_contacts(
        df if df is not None else pd.DataFrame(columns=SITE_CONTACT_COLUMNS)
    )


def save_site_registry(df, path=None, create_backup=True, df_contacts=None):
    path = Path(path) if path else caminho_sites_excel()
    path.parent.mkdir(parents=True, exist_ok=True)

    contacts_save = prepare_contacts_for_save(
        df_contacts if df_contacts is not None else load_site_contacts(path)
    )
    backup = backup_site_registry(path) if create_backup else None
    df_save = prepare_registry_for_save(df)

    workbook = load_or_create_workbook(path)
    base_sheet = workbook[registry_sheet_name(workbook)]

    write_dataframe_preserving_sheet(
        base_sheet,
        df_save,
        SITE_REGISTRY_COLUMNS,
        COLUMN_ALIASES
    )

    if SITE_CONTACTS_SHEET in workbook.sheetnames:
        contacts_sheet = workbook[SITE_CONTACTS_SHEET]
    else:
        contacts_sheet = workbook.create_sheet(SITE_CONTACTS_SHEET)

    write_dataframe_preserving_sheet(
        contacts_sheet,
        contacts_save,
        SITE_CONTACT_COLUMNS,
        CONTACT_COLUMN_ALIASES
    )

    workbook.save(path)

    return backup


def save_site_contacts(df_contacts, path=None, create_backup=True):
    df_sites = load_site_registry(path)

    return save_site_registry(
        df_sites,
        path=path,
        create_backup=create_backup,
        df_contacts=df_contacts
    )


def duplicated_site_codes(df):
    if df is None or df.empty:
        return pd.DataFrame(
            columns=[
                "Campo",
                "Código",
                "Quantidade",
                "Código Aquiles",
                "Código Microsiga",
                "Código Condomínio",
                "SNMPc",
                "Nome",
                "Status"
            ]
        )

    df_base = df.copy()

    for column in SITE_REGISTRY_COLUMNS:
        if column not in df_base.columns:
            df_base[column] = ""

    registros = []

    for column in UNIQUE_SITE_CODE_COLUMNS:
        codigos = df_base[column].apply(normalize_code)
        duplicados = codigos[
            codigos.ne("")
            & codigos.duplicated(keep=False)
        ]

        for codigo in sorted(duplicados.unique()):
            linhas = df_base[codigos == codigo]

            for _, row in linhas.iterrows():
                registros.append({
                    "Campo": UNIQUE_SITE_CODE_LABELS.get(column, column),
                    "Código": codigo,
                    "Quantidade": len(linhas),
                    "Código Aquiles": normalize_code(row.get("CÓDIGO AQUILES")),
                    "Código Microsiga": normalize_code(row.get("CÓDIGO MICROSIGA")),
                    "Código Condomínio": normalize_code(row.get("CÓDIGO CONDOMINIO")),
                    "SNMPc": row.get("SMNPC", ""),
                    "Nome": row.get("NOME", ""),
                    "Status": row.get("Status", "")
                })

    return pd.DataFrame(registros)


def validate_unique_site_codes(df, record, original_code=None):
    original_code = normalize_code(original_code)

    if df is None:
        df = pd.DataFrame(columns=SITE_REGISTRY_COLUMNS)

    df = df.copy()

    for column in SITE_REGISTRY_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    for column in UNIQUE_SITE_CODE_COLUMNS:
        code = normalize_code(record.get(column))

        if not code:
            continue

        codes = df[column].apply(normalize_code)
        duplicates = df[
            codes.eq(code)
            & df[SITE_CODE_COLUMN].apply(normalize_code).ne(original_code)
        ]

        if duplicates.empty:
            continue

        label = UNIQUE_SITE_CODE_LABELS.get(column, column)
        nomes = ", ".join(
            str(valor).strip()
            for valor in duplicates.get("SMNPC", pd.Series(dtype=object)).tolist()
            if str(valor).strip()
        )
        detalhe = f" em {nomes}" if nomes else ""

        raise ValueError(
            f"Ja existe outro site com {label} {code}{detalhe}."
        )


def upsert_site(record, original_code=None):
    df = load_site_registry()
    code = normalize_code(record.get(SITE_CODE_COLUMN))
    original_code = normalize_code(original_code)

    if not code:
        raise ValueError("Informe o codigo do site.")

    validate_unique_site_codes(
        df,
        record,
        original_code=original_code
    )

    codes = df[SITE_CODE_COLUMN].apply(normalize_code)

    if original_code and original_code in set(codes):
        index = codes[codes == original_code].index[0]

        for column in SITE_REGISTRY_COLUMNS:
            df.at[index, column] = record.get(column, "")

    else:
        df = pd.concat(
            [
                df,
                pd.DataFrame([
                    {
                        column: record.get(column, "")
                        for column in SITE_REGISTRY_COLUMNS
                    }
                ])
            ],
            ignore_index=True
        )

    return save_site_registry(df)


def export_site_registry_excel(df=None):
    df_export = prepare_registry_for_save(
        df if df is not None else load_site_registry()
    )
    df_contacts = prepare_contacts_for_save(
        load_site_contacts()
    )
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(
            writer,
            index=False,
            sheet_name="BASE"
        )
        df_contacts.to_excel(
            writer,
            index=False,
            sheet_name=SITE_CONTACTS_SHEET
        )

    output.seek(0)

    return output.getvalue()
