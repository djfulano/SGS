import tempfile
import unittest
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from app.models.cliente import Cliente
from app.models.site import Site
from app.services import finance_service as fs


class FinancePrioritiesTest(unittest.TestCase):

    def cadastro(self):
        return pd.DataFrame([
            {
                "CÓDIGO AQUILES": "100",
                "CÓDIGO MICROSIGA": "12345",
                "SMNPC": "SITE_A_POP_100_IP",
                "NOME": "SITE A",
                "Status": "Ativo",
                "LOCAÇÃO": 900,
                "DIA VENCIMENTO": 15,
                "SITE CRÍTICO": "Sim",
                "TIPO CRITICIDADE": "Bloqueia",
            },
            {
                "CÓDIGO AQUILES": "200",
                "CÓDIGO MICROSIGA": "23456",
                "SMNPC": "SITE_B_BH_200_IP",
                "NOME": "SITE B",
                "Status": " ativo ",
                "LOCAÇÃO": 700,
                "DIA VENCIMENTO": 10,
                "SITE CRÍTICO": "Não",
                "TIPO CRITICIDADE": "",
            },
            {
                "CÓDIGO AQUILES": "300",
                "CÓDIGO MICROSIGA": "34567",
                "SMNPC": "SITE_C_REP_300_IP",
                "NOME": "SITE C",
                "Status": "Cancelado",
                "LOCAÇÃO": 500,
                "DIA VENCIMENTO": 5,
                "SITE CRÍTICO": "Não",
                "TIPO CRITICIDADE": "",
            },
        ])

    def pagamentos(self):
        return pd.DataFrame([
            {
                "ID SGS": "R1",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "RECORRENTE",
                "Data de vencimento": "2026-07-01",
                "Subtotal": 100.0,
                "Site localizado": "Sim",
            },
            {
                "ID SGS": "R2",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "RECORRENTE",
                "Data de vencimento": "2026-08-01",
                "Subtotal": 120.0,
                "Site localizado": "Sim",
            },
            {
                "ID SGS": "A1",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "ACORDO/PARCELAMENTO",
                "Data de vencimento": "2026-06-01",
                "Subtotal": 50.0,
                "Site localizado": "Sim",
            },
            {
                "ID SGS": "A2",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "ACORDO/PARCELAMENTO",
                "Data de vencimento": "2026-09-01",
                "Subtotal": 60.0,
                "Site localizado": "Sim",
            },
            {
                "ID SGS": "R3",
                "Status": "Pago",
                "Microsiga": "012345",
                "Tipo de despesa": "RECORRENTE",
                "Data de vencimento": "2026-05-01",
                "Subtotal": 999.0,
                "Site localizado": "Sim",
            },
        ])

    def sites_topologia(self):
        principal = Site("SITE_A_POP_100_IP", "POP")
        principal.codigo_topos = "100"
        principal.microsiga = "12345"
        principal.adicionar_cliente(
            Cliente("Cliente Principal", 500, "10000001")
        )
        filho = Site("SITE_A_BH_101_IP", "BH")
        filho.adicionar_cliente(
            Cliente("Cliente Filho", 300, "10000002")
        )
        principal.adicionar_filho(filho)
        return {
            principal.nome: principal,
            filho.nome: filho,
        }

    def test_monta_lista_com_parcelas_e_fallback_do_cadastro(self):
        resultado = fs.montar_prioridades_financeiras(
            cadastro_sites=self.cadastro(),
            pagamentos=self.pagamentos(),
            prioridades={
                "sites": {
                    "aquiles:100": {"importance": "Alta"},
                }
            },
            hoje=date(2026, 7, 31),
            sites=self.sites_topologia(),
        )

        self.assertEqual(len(resultado), 2)
        site_a = resultado.set_index("Chave Site").loc["aquiles:100"]
        self.assertEqual(
            site_a["Site"],
            "SITE_A_POP_100_IP - 100 / SITE A - 12345",
        )
        self.assertEqual(site_a["Vencimento da Mensalidade"], date(2026, 8, 1))
        self.assertEqual(site_a["Valor da Mensalidade Atual"], 120.0)
        self.assertEqual(site_a["Tem Acordo"], "Sim")
        self.assertEqual(site_a["Vencimento do Acordo"], date(2026, 6, 1))
        self.assertEqual(site_a["Valor da Parcela do Acordo"], 50.0)
        self.assertEqual(site_a["Mensalidades Vencidas"], 1)
        self.assertEqual(site_a["Valor das Mensalidades Vencidas"], 100.0)
        self.assertEqual(site_a["Acordos Vencidos"], 1)
        self.assertEqual(site_a["Valor dos Acordos Vencidos"], 50.0)
        self.assertEqual(site_a["Total de Mensalidades em Atraso"], 100.0)
        self.assertEqual(site_a["Quantidade de Mensalidades em Atraso"], 1)
        self.assertEqual(site_a["Valor Médio das Mensalidades em Atraso"], 100.0)
        self.assertEqual(site_a["Total de Acordos Atrasados"], 50.0)
        self.assertEqual(
            site_a["Quantidade de Parcelas de Acordos Atrasados"],
            1,
        )
        self.assertEqual(site_a["Média de Valores de Acordos Atrasados"], 50.0)
        self.assertEqual(site_a["Valor Total em Atraso"], 150.0)
        self.assertEqual(site_a["Criticidade"], "Bloqueia")
        self.assertEqual(site_a["Importância"], "Alta")
        self.assertEqual(site_a["Microsiga"], "012345")
        self.assertEqual(site_a["Custo mensal"], 900.0)
        self.assertEqual(site_a["Receita (Total com sites filhos)"], 800.0)
        self.assertEqual(site_a["Passivo de mensalidades"], 220.0)
        self.assertEqual(site_a["Passivo de acordos"], 110.0)
        self.assertEqual(
            site_a["Data Vencimento Mensalidade"],
            date(2026, 8, 1),
        )
        self.assertEqual(
            site_a["Data Vencimento Acordo"],
            date(2026, 6, 1),
        )
        self.assertEqual(site_a["Quantidade de parcelas em atraso"], 2)
        self.assertEqual(
            site_a["Lista de Clientes"],
            "Cliente Filho, Cliente Principal",
        )

        site_b = resultado.set_index("Chave Site").loc["aquiles:200"]
        self.assertEqual(site_b["Vencimento da Mensalidade"], date(2026, 8, 10))
        self.assertEqual(site_b["Valor da Mensalidade Atual"], 700.0)
        self.assertEqual(site_b["Tem Acordo"], "Não")
        self.assertEqual(site_b["Criticidade"], "Não crítico")
        self.assertEqual(site_b["Importância"], "Não definida")

    def test_lista_sem_pagamentos_mantem_sites_ativos(self):
        resultado = fs.montar_prioridades_financeiras(
            cadastro_sites=self.cadastro(),
            pagamentos=pd.DataFrame(columns=fs.PAYMENT_COLUMNS),
            prioridades={"sites": {}},
            hoje=date(2026, 7, 31),
        )
        self.assertEqual(set(resultado["Chave Site"]), {"aquiles:100", "aquiles:200"})
        self.assertEqual(int(resultado["Mensalidades Vencidas"].sum()), 0)
        self.assertEqual(int(resultado["Acordos Vencidos"].sum()), 0)
        self.assertEqual(float(resultado["Valor Médio das Mensalidades em Atraso"].sum()), 0.0)
        self.assertEqual(float(resultado["Média de Valores de Acordos Atrasados"].sum()), 0.0)
        self.assertEqual(float(resultado["Valor Total em Atraso"].sum()), 0.0)

    def test_mensalidade_atual_usa_vencida_mais_recente_sem_parcela_futura(self):
        pagamentos = pd.DataFrame([
            {
                "ID SGS": "R1",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "RECORRENTE",
                "Data de vencimento": "2026-05-01",
                "Subtotal": 100.0,
                "Site localizado": "Sim",
            },
            {
                "ID SGS": "R2",
                "Status": "Pendente",
                "Microsiga": "012345",
                "Tipo de despesa": "RECORRENTE",
                "Data de vencimento": "2026-07-01",
                "Subtotal": 130.0,
                "Site localizado": "Sim",
            },
        ])
        resultado = fs.montar_prioridades_financeiras(
            cadastro_sites=self.cadastro(),
            pagamentos=pagamentos,
            prioridades={"sites": {}},
            hoje=date(2026, 8, 11),
        )
        site_a = resultado.set_index("Chave Site").loc["aquiles:100"]
        self.assertEqual(site_a["Data Vencimento Mensalidade"], date(2026, 7, 1))
        self.assertEqual(site_a["Valor da Mensalidade Atual"], 130.0)

    def test_salva_importancia_com_auditoria_e_recarrega(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "site_priorities.json"
            with patch.object(fs, "registrar_log_sistema") as registrar:
                salvo = fs.salvar_importancias_sites(
                    [{"site_key": "aquiles:100", "importance": "Crítica"}],
                    usuario="financeiro",
                    path=path,
                    agora=datetime(2026, 7, 31, 10, 30),
                )
            self.assertEqual(salvo["alteracoes"], 1)
            registro = fs.carregar_prioridades_sites(path)["sites"]["aquiles:100"]
            self.assertEqual(registro["importance"], "Crítica")
            self.assertEqual(registro["updated_by"], "financeiro")
            self.assertEqual(registro["updated_at"], "2026-07-31T10:30:00")
            registrar.assert_called_once()

    def test_recusa_importancia_invalida(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaisesRegex(ValueError, "Importância inválida"):
                fs.salvar_importancias_sites(
                    [{"site_key": "aquiles:100", "importance": "Urgente"}],
                    path=Path(temp_dir) / "site_priorities.json",
                )

    def test_exporta_valores_numericos_com_formato_em_reais(self):
        dados = pd.DataFrame([{
            "Chave Site": "aquiles:100",
            "Site": "SITE A - 100 / SITE A - 12345",
            "Microsiga": "012345",
            "Custo mensal": 1234.56,
            "Receita (Total com sites filhos)": 2000.0,
            "Data Vencimento Mensalidade": date(2026, 8, 10),
            "Valor da Mensalidade Atual": 900.0,
            "Total de Mensalidades em Atraso": 300.0,
            "Quantidade de Mensalidades em Atraso": 3,
            "Valor Médio das Mensalidades em Atraso": 100.0,
            "Data Vencimento Acordo": date(2026, 8, 20),
            "Total de Acordos Atrasados": 150.0,
            "Quantidade de Parcelas de Acordos Atrasados": 2,
            "Média de Valores de Acordos Atrasados": 75.0,
            "Valor Total em Atraso": 450.0,
            "Criticidade": "Bloqueia",
            "Importância": "Alta",
            "Lista de Clientes": "Cliente A, Cliente B",
            "Possui Vencidos": "Sim",
        }])
        arquivo = fs.exportar_prioridades_financeiras_excel(dados)
        planilha = load_workbook(BytesIO(arquivo))["Prioridades"]
        cabecalhos = {
            celula.value: celula.column
            for celula in planilha[1]
        }
        custo = planilha.cell(
            row=2,
            column=cabecalhos["Custo mensal"],
        )
        acordos = planilha.cell(
            row=2,
            column=cabecalhos["Total de Acordos Atrasados"],
        )
        self.assertEqual(custo.data_type, "n")
        self.assertEqual(custo.value, 1234.56)
        self.assertIn("R$", custo.number_format)
        self.assertEqual(acordos.data_type, "n")
        self.assertIn("R$", acordos.number_format)
        self.assertIn(
            "R$",
            planilha.cell(
                row=2,
                column=cabecalhos["Valor Total em Atraso"],
            ).number_format,
        )
        vencimento_mensalidade = planilha.cell(
            row=2,
            column=cabecalhos["Data Vencimento Mensalidade"],
        )
        vencimento_acordo = planilha.cell(
            row=2,
            column=cabecalhos["Data Vencimento Acordo"],
        )
        self.assertEqual(vencimento_mensalidade.value.date(), date(2026, 8, 10))
        self.assertEqual(vencimento_acordo.value.date(), date(2026, 8, 20))
        self.assertEqual(vencimento_mensalidade.number_format, "DD/MM/YYYY")
        self.assertEqual(vencimento_acordo.number_format, "DD/MM/YYYY")
        self.assertEqual(
            planilha.cell(
                row=2,
                column=cabecalhos["Quantidade de Mensalidades em Atraso"],
            ).value,
            3,
        )
        self.assertEqual(
            planilha.cell(
                row=2,
                column=cabecalhos["Quantidade de Parcelas de Acordos Atrasados"],
            ).value,
            2,
        )
        self.assertEqual(
            list(cabecalhos),
            fs.SITE_PRIORITY_EXPORT_COLUMNS,
        )
        self.assertNotIn("Chave Site", cabecalhos)
        self.assertNotIn("Possui Vencidos", cabecalhos)
        self.assertNotIn("Passivo de acordos", cabecalhos)
        self.assertNotIn("Passivo de mensalidades", cabecalhos)

    def test_exportacao_preserva_valor_restrito_como_texto(self):
        dados = pd.DataFrame([{
            "Site": "SITE A",
            "Custo mensal": "Restrito",
        }])
        arquivo = fs.exportar_prioridades_financeiras_excel(dados)
        planilha = load_workbook(BytesIO(arquivo))["Prioridades"]
        cabecalhos = {
            celula.value: celula.column
            for celula in planilha[1]
        }
        valor = planilha.cell(
            row=2,
            column=cabecalhos["Custo mensal"],
        )
        self.assertEqual(valor.value, "Restrito")
        self.assertEqual(valor.data_type, "s")


if __name__ == "__main__":
    unittest.main()
