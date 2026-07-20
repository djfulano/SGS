import tempfile
import unittest
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook
import pandas as pd

from app.services import finance_service as fs


class FinanceServiceTest(unittest.TestCase):

    def setUp(self):
        patcher = patch.object(
            fs,
            "load_site_registry",
            return_value=pd.DataFrame(),
        )
        patcher.start()
        self.addCleanup(patcher.stop)

    def criar_topos(self, linhas=None):
        linhas = linhas or [
            {
                "ANO": 2023,
                "COMPETÊNCIA MÊS DO VENCIMENTO": "2023-01-01",
                "STATUS (A VENCER/PAGO/ EM ATRASO)": "A VENCER",
                "APROVAÇÃO/NEGOCIAÇÃO": "APROVADO",
                "O.C / CONTA CONTÁBIL": "OC-1",
                "DATA PAGAMENTO (FLUXO DE CAIXA)": "2026-08-10",
                "VENCIMENTO ORIGINAL": "2023-01-15",
                "PRIORIDADE": "ALTA",
                "TIPO PGTO": "PIX",
                "FAVORECIDO": "SITE TESTE 092159",
                "VALOR": 900,
                "JUROS/DESC/CRED/DIF A PAGAR": 100,
                "VALOR TOTAL A PAGAR": 1000,
                "TIPO DE DESPESA": "RECORRENTE",
                "DESCRIÇÃO": "LOCAÇÃO",
            },
            {
                "ANO": 2026,
                "COMPETÊNCIA MÊS DO VENCIMENTO": "2026-08-01",
                "STATUS (A VENCER/PAGO/ EM ATRASO)": "EM ATRASO",
                "O.C / CONTA CONTÁBIL": "OC-2",
                "DATA PAGAMENTO (FLUXO DE CAIXA)": "2026-08-20",
                "VENCIMENTO ORIGINAL": "2026-08-10",
                "PRIORIDADE": "MÉDIA",
                "FAVORECIDO": "SITE TESTE 092159",
                "VALOR TOTAL A PAGAR": 500,
                "TIPO DE DESPESA": "ACORDO/PARCELAMENTO",
                "DESCRIÇÃO": "PARCELA ACORDO",
            },
        ]
        path = Path(tempfile.mkstemp(suffix=".xlsx")[1])
        pd.DataFrame(linhas).to_excel(path, index=False, sheet_name="TOPOS")
        return path

    def criar_planilha(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Fechamento 2016 a 2025"
        headers = [
            "Ano", "Prioridade", "CNPJ/CPF", "Tipo", "PIX", "Banco", "Cód", "Agência",
            "C/Corrente", "Multa", "Juros", "Mês", "Nome", "Fornecedor", "Microsiga",
            "Produto", "C.C", "RC NOVA", "OC NOVA ", "OC Primário", "OC Secundário",
            "Vencto", "Energia", "Outros", "Locação", "Subtotal", "Descrição"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=6, column=col, value=header)
        row = [
            2026, "Prioridade", "00.000.000/0001-00", "TED", "", "ITAU", "341", "1234",
            "12345-6", 0.02, 0.01, 46023, "SITE TESTE", "FORNECEDOR TESTE", 92159,
            "", "", "RC1", "OCN", "OCP", "OCS", 18, 100, 20, 1000, 1120, "Pagamento teste"
        ]
        for col, value in enumerate(row, start=1):
            ws.cell(row=7, column=col, value=value)
        row_antigo = row.copy()
        row_antigo[0] = 2025
        row_antigo[12] = "SITE ANTIGO"
        row_antigo[13] = "FORNECEDOR ANTIGO"
        row_antigo[25] = 999
        for col, value in enumerate(row_antigo, start=1):
            ws.cell(row=8, column=col, value=value)

        ws2 = wb.create_sheet("Acordos")
        headers2 = ["", "Obs", "Mês", "Acordo", "Nome", "Resp", "Aprovado Sindico", "PAGO", "Microsiga", "Acordo2", "Descrição", "", "", "", "", "", "Multa + Juros"]
        for col, header in enumerate(headers2, start=1):
            ws2.cell(row=9, column=col, value=header)
        row2 = ["", "Obs teste", 46023, "Documento", "SITE TESTE", "Financeiro", "SIM", "", 92159, 500, "Acordo teste", "", "", "", "", "", 50]
        for col, value in enumerate(row2, start=1):
            ws2.cell(row=10, column=col, value=value)
        row2_antigo = row2.copy()
        row2_antigo[2] = 45658
        row2_antigo[4] = "ACORDO ANTIGO"
        for col, value in enumerate(row2_antigo, start=1):
            ws2.cell(row=11, column=col, value=value)

        ws3 = wb.create_sheet("Acordos Gráficos")
        ws3["A1"] = "Ignorar"

        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm")
        wb.save(temp.name)
        temp.close()
        return Path(temp.name)

    def test_importa_pagamentos_e_acordos_com_vinculo(self):
        path = self.criar_planilha()
        sites = {
            "SITE_SNMP": SimpleNamespace(
                microsiga="092159",
                codigo_topos="654321",
                nome="SITE_SNMP",
                nome_cadastro="SITE CADASTRO",
            )
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            payments = Path(temp_dir) / "payments.json"
            agreements = Path(temp_dir) / "agreements.json"
            with patch.object(fs, "PAYMENTS_FILE", payments), patch.object(fs, "AGREEMENTS_FILE", agreements):
                resultado = fs.importar_planilha_financeira(path, sites=sites, salvar=True, usuario="teste")
                self.assertEqual(resultado["resumo"]["pagamentos"]["importados"], 1)
                self.assertEqual(resultado["resumo"]["acordos"]["importados"], 1)
                self.assertEqual(resultado["resumo"]["pagamentos"]["com_site"], 1)
                self.assertEqual(resultado["resumo"]["pagamentos"]["sem_site"], 0)
                self.assertNotIn("SITE ANTIGO", set(resultado["pagamentos"]["Nome"]))
                self.assertNotIn("ACORDO ANTIGO", set(resultado["acordos"]["Nome"]))
                self.assertEqual(resultado["pagamentos"].iloc[0]["Nome SNMPc"], "SITE_SNMP")
                self.assertEqual(resultado["pagamentos"].iloc[0]["Microsiga"], "092159")
                self.assertEqual(resultado["pagamentos"].iloc[0]["OC NOVA"], "OCN")
                self.assertEqual(resultado["pagamentos"].iloc[0]["Subtotal"], 1120.0)
                self.assertEqual(resultado["acordos"].iloc[0]["Status"], "Aprovado")
                self.assertEqual(resultado["acordos"].iloc[0]["Valor Acordo"], 500.0)

                segunda = fs.importar_planilha_financeira(path, sites=sites, salvar=False, usuario="teste")
                self.assertEqual(segunda["resumo"]["pagamentos"]["novos"], 0)
                self.assertGreaterEqual(segunda["resumo"]["pagamentos"]["duplicados"], 1)

    def test_status_pagamento_vencido_derivado(self):
        row = {"Status": "Pendente", "Data de vencimento": "2026-01-01"}
        self.assertEqual(fs.status_pagamento_exibicao(row), "Vencido")
        row_pago = {"Status": "Pago", "Data de vencimento": "2026-01-01"}
        self.assertEqual(fs.status_pagamento_exibicao(row_pago), "Pago")

    def test_importa_topos_todos_os_anos_e_classifica_acordo(self):
        path = self.criar_topos()
        pagamentos, acordos = fs.ler_topos_em_aberto_excel(path)
        self.assertEqual(len(pagamentos), 2)
        self.assertEqual(len(acordos), 1)
        self.assertEqual(pagamentos.iloc[0]["Microsiga"], "092159")
        self.assertEqual(pagamentos.iloc[0]["Subtotal"], 1000.0)
        self.assertEqual(pagamentos.iloc[0]["Data de vencimento"], "2023-01-15")
        self.assertEqual(acordos.iloc[0]["Tipo de despesa"], "ACORDO/PARCELAMENTO")
        self.assertEqual(acordos.iloc[0]["Valor Acordo"], 500.0)

    def test_topos_preserva_linhas_identicas_com_ids_distintos(self):
        linha = {
            "FAVORECIDO": "SITE 092159",
            "VENCIMENTO ORIGINAL": "2026-01-10",
            "VALOR TOTAL A PAGAR": 100,
            "TIPO DE DESPESA": "RECORRENTE",
            "O.C / CONTA CONTÁBIL": "OC",
        }
        pagamentos, _ = fs.ler_topos_em_aberto_excel(self.criar_topos([linha, linha.copy()]))
        self.assertEqual(len(pagamentos), 2)
        self.assertEqual(pagamentos["ID SGS"].nunique(), 2)

    def test_primeira_importacao_topos_exige_confirmacao_para_substituir(self):
        path = self.criar_topos()
        sites = {
            "SITE_SNMP": SimpleNamespace(
                microsiga="092159",
                codigo_topos="654321",
                nome="SITE_SNMP",
                nome_cadastro="SITE CADASTRO",
            )
        }
        antigo = pd.DataFrame([{**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "PAG-ANTIGO", "Status": "Pendente"}])
        with tempfile.TemporaryDirectory() as temp_dir:
            payments = Path(temp_dir) / "payments.json"
            agreements = Path(temp_dir) / "agreements.json"
            with patch.object(fs, "PAYMENTS_FILE", payments), patch.object(fs, "AGREEMENTS_FILE", agreements):
                fs.salvar_pagamentos(antigo)
                previa = fs.importar_planilha_financeira(path, sites=sites)
                self.assertTrue(previa["requer_substituicao_base_antiga"])
                with self.assertRaises(ValueError):
                    fs.importar_planilha_financeira(path, sites=sites, salvar=True)
                fs.importar_planilha_financeira(
                    path,
                    sites=sites,
                    salvar=True,
                    substituir_base_antiga=True,
                )
                salvos = fs.carregar_pagamentos()
                self.assertNotIn("PAG-ANTIGO", set(salvos["ID SGS"]))
                self.assertEqual(len(salvos), 2)

    def test_dashboard_topos_calcula_atraso_acordos_e_sites(self):
        pagamentos = pd.DataFrame([
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "1", "Status": "Pendente", "Data de vencimento": "2026-01-01", "Data programada pagamento": "2026-07-20", "Subtotal": 1000.0, "Microsiga": "000001", "Site localizado": "Sim", "Nome Site": "A", "Nome SNMPc": "A"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "2", "Status": "Pendente", "Data de vencimento": "2026-02-01", "Subtotal": 500.0, "Microsiga": "000002", "Site localizado": "Sim", "Nome Site": "B", "Nome SNMPc": "B"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "3", "Status": "Pendente", "Data de vencimento": "2026-08-10", "Data programada pagamento": "2026-07-17", "Tipo de despesa": "RECORRENTE", "Subtotal": 300.0, "Microsiga": "000003", "Site localizado": "Sim", "Nome Site": "C", "Nome SNMPc": "C"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "4", "Status": "Pendente", "Data de vencimento": "2026-08-20", "Tipo de despesa": "ACORDO/PARCELAMENTO", "Subtotal": 200.0, "Microsiga": "000004", "Site localizado": "Sim", "Nome Site": "D", "Nome SNMPc": "D"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "5", "Status": "Pago", "Data de vencimento": "2026-08-20", "Tipo de despesa": "RECORRENTE", "Subtotal": 900.0, "Microsiga": "000005", "Site localizado": "Sim", "Nome Site": "E", "Nome SNMPc": "E"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "6", "Status": "Cancelado", "Data de vencimento": "2026-08-20", "Tipo de despesa": "ACORDO/PARCELAMENTO", "Subtotal": 800.0, "Microsiga": "000006", "Site localizado": "Sim", "Nome Site": "F", "Nome SNMPc": "F"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "7", "Status": "Pendente", "Data de vencimento": "2026-07-10", "Tipo de despesa": "RECORRENTE", "Subtotal": 75.0, "Microsiga": "000002", "Site localizado": "Sim", "Nome Site": "B", "Nome SNMPc": "B"},
        ])
        acordos = pd.DataFrame([
            {**{c: "" for c in fs.AGREEMENT_COLUMNS}, "ID SGS": "A1", "Status": "Em pagamento", "Valor Acordo": 1000.0, "Microsiga": "000001", "Site localizado": "Sim"},
        ])
        with patch.object(fs, "carregar_pagamentos", return_value=pagamentos), patch.object(fs, "carregar_acordos", return_value=acordos), patch.object(fs, "date") as date_mock:
            date_mock.today.return_value = date(2026, 7, 17)
            dados = fs.dashboard_financeiro(hoje=date(2026, 7, 17))
        self.assertEqual(dados["total_vencido"], 1575.0)
        self.assertEqual(dados["total_acordos_abertos"], 1000.0)
        self.assertEqual(dados["sites_com_acordo"], 1)
        self.assertEqual(dados["sites_atrasados_sem_acordo"], 1)
        self.assertEqual(dados["programacao_mensal"].to_dict(orient="records"), [
            {"Mês": "2026-07", "Mensalidades": 75.0, "Acordos": 0.0, "Total": 75.0},
            {"Mês": "2026-08", "Mensalidades": 300.0, "Acordos": 200.0, "Total": 500.0},
        ])
        self.assertEqual(float(dados["origem_mensal"]["Valor"].sum()), 1575.0)
        self.assertNotIn("2026-08", set(dados["origem_mensal"]["Mês"]))

    def test_programacao_mensal_mantem_meses_com_uma_categoria(self):
        pagamentos = pd.DataFrame([
            {"Data de vencimento": "2026-09-10", "Tipo de despesa": "RECORRENTE", "Subtotal": 100.0},
            {"Data de vencimento": "2026-10-10", "Tipo de despesa": "ACORDO/PARCELAMENTO", "Subtotal": 50.0},
        ])
        serie = fs._programacao_mensal_por_tipo(pagamentos)
        self.assertEqual(serie.to_dict(orient="records"), [
            {"Mês": "2026-09", "Mensalidades": 100.0, "Acordos": 0.0, "Total": 100.0},
            {"Mês": "2026-10", "Mensalidades": 0.0, "Acordos": 50.0, "Total": 50.0},
        ])

    def test_programacao_mensal_inicia_no_mes_atual_e_preenche_intervalos(self):
        pagamentos = pd.DataFrame([
            {"Data de vencimento": "2026-09-10", "Tipo de despesa": "RECORRENTE", "Subtotal": 100.0},
        ])
        serie = fs._programacao_mensal_por_tipo(pagamentos, inicio=date(2026, 7, 1))
        self.assertEqual(serie.to_dict(orient="records"), [
            {"Mês": "2026-07", "Mensalidades": 0.0, "Acordos": 0.0, "Total": 0.0},
            {"Mês": "2026-08", "Mensalidades": 0.0, "Acordos": 0.0, "Total": 0.0},
            {"Mês": "2026-09", "Mensalidades": 100.0, "Acordos": 0.0, "Total": 100.0},
        ])

    def test_reimportacao_topos_preserva_ausentes_status_e_observacao(self):
        path = self.criar_topos()
        sites = {
            "SITE_SNMP": SimpleNamespace(
                microsiga="092159",
                codigo_topos="654321",
                nome="SITE_SNMP",
                nome_cadastro="SITE CADASTRO",
            )
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            payments = Path(temp_dir) / "payments.json"
            agreements = Path(temp_dir) / "agreements.json"
            with patch.object(fs, "PAYMENTS_FILE", payments), patch.object(fs, "AGREEMENTS_FILE", agreements):
                fs.importar_planilha_financeira(path, sites=sites, salvar=True)
                base = fs.carregar_pagamentos()
                primeiro_id = base.iloc[0]["ID SGS"]
                base.loc[base["ID SGS"].eq(primeiro_id), "Status"] = "Pago"
                base.loc[base["ID SGS"].eq(primeiro_id), "Observação interna"] = "Baixa manual"
                fs.salvar_pagamentos(base)

                uma_linha = self.criar_topos([{
                    "ANO": 2023,
                    "COMPETÊNCIA MÊS DO VENCIMENTO": "2023-01-01",
                    "O.C / CONTA CONTÁBIL": "OC-1",
                    "DATA PAGAMENTO (FLUXO DE CAIXA)": "2026-08-10",
                    "VENCIMENTO ORIGINAL": "2023-01-15",
                    "FAVORECIDO": "SITE TESTE 092159",
                    "VALOR": 900,
                    "JUROS/DESC/CRED/DIF A PAGAR": 100,
                    "VALOR TOTAL A PAGAR": 1000,
                    "TIPO DE DESPESA": "RECORRENTE",
                    "DESCRIÇÃO": "LOCAÇÃO",
                }])
                fs.importar_planilha_financeira(uma_linha, sites=sites, salvar=True)
                recarregada = fs.carregar_pagamentos()
                self.assertEqual(len(recarregada), 2)
                preservado = recarregada[recarregada["ID SGS"].eq(primeiro_id)].iloc[0]
                self.assertEqual(preservado["Status"], "Pago")
                self.assertEqual(preservado["Observação interna"], "Baixa manual")

    def test_conciliacao_identifica_problemas_sem_alterar_bases(self):
        sites = {
            "ATIVO": SimpleNamespace(
                microsiga="1", codigo_topos="101", nome="ATIVO",
                nome_cadastro="Site Ativo", status_cadastro="Ativo",
            ),
            "INATIVO": SimpleNamespace(
                microsiga="2", codigo_topos="102", nome="INATIVO",
                nome_cadastro="Site Inativo", status_cadastro="Cancelado",
            ),
            "DUP_A": SimpleNamespace(
                microsiga="3", codigo_topos="103", nome="DUP_A",
                nome_cadastro="Duplicado A", status_cadastro="Ativo",
            ),
            "DUP_B": SimpleNamespace(
                microsiga="000003", codigo_topos="104", nome="DUP_B",
                nome_cadastro="Duplicado B", status_cadastro="Ativo",
            ),
        }
        pagamentos = pd.DataFrame([
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P1", "Status": "Pendente", "Favorecido": "SEM CODIGO", "Data de vencimento": "2026-08-01", "Subtotal": 100.0, "Site localizado": "Não"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P2", "Status": "Pendente", "Favorecido": "DESCONHECIDO 999999", "Data de vencimento": "2026-08-01", "Subtotal": 200.0, "Site localizado": "Não"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P3", "Status": "Pendente", "Microsiga": "1", "Nome SNMPc": "ANTIGO", "Código Aquiles": "999", "Data de vencimento": "2026-08-01", "Subtotal": 300.0, "Site localizado": "Sim"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P4", "Status": "Pendente", "Microsiga": "2", "Nome SNMPc": "INATIVO", "Data de vencimento": "2026-08-01", "Subtotal": 400.0, "Site localizado": "Sim"},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P5", "Status": "Pendente", "Microsiga": "1", "Data de vencimento": "inválido", "Subtotal": 0.0, "Site localizado": "Não"},
        ])
        acordos = pd.DataFrame([{
            **{c: "" for c in fs.AGREEMENT_COLUMNS},
            "ID SGS": "A1", "ID Pagamento": "INEXISTENTE",
            "Status": "Em pagamento", "Microsiga": "1",
            "Data de vencimento": "2026-08-01", "Valor Acordo": 500.0,
            "Nome SNMPc": "ATIVO", "Site localizado": "Sim",
        }])
        pagamentos_antes = pagamentos.copy(deep=True)
        acordos_antes = acordos.copy(deep=True)

        resultado = fs.analisar_conciliacao_financeira(
            sites, pagamentos=pagamentos, acordos=acordos
        )

        tipos = set(resultado["Tipo do problema"])
        self.assertTrue({
            "Sem Código Microsiga",
            "Código Microsiga não encontrado",
            "Código Microsiga duplicado no cadastro",
            "Vínculo incompatível com o cadastro atual",
            "Vínculo com site inativo",
            "Acordo sem pagamento de origem",
            "Vencimento ausente ou inválido",
            "Valor ausente, zero ou inválido",
        }.issubset(tipos))
        pd.testing.assert_frame_equal(pagamentos, pagamentos_antes)
        pd.testing.assert_frame_equal(acordos, acordos_antes)

    def test_historico_financeiro_site_separa_status_e_valores(self):
        pagamentos = pd.DataFrame([
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "V", "Status": "Pendente", "Microsiga": "1", "Data de vencimento": "2026-06-01", "Subtotal": 100.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "F", "Status": "Pendente", "Microsiga": "000001", "Data de vencimento": "2026-08-01", "Subtotal": 200.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P", "Status": "Pago", "Microsiga": "1", "Data de vencimento": "2026-05-01", "Subtotal": 300.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "C", "Status": "Cancelado", "Microsiga": "1", "Data de vencimento": "2026-04-01", "Subtotal": 400.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "OUTRO", "Status": "Pendente", "Microsiga": "2", "Data de vencimento": "2026-06-01", "Subtotal": 900.0},
        ])
        acordos = pd.DataFrame([
            {**{c: "" for c in fs.AGREEMENT_COLUMNS}, "ID SGS": "A1", "Status": "Em pagamento", "Microsiga": "1", "Valor Acordo": 500.0},
            {**{c: "" for c in fs.AGREEMENT_COLUMNS}, "ID SGS": "A2", "Status": "Quitado", "Microsiga": "1", "Valor Acordo": 600.0},
        ])

        historico = fs.historico_financeiro_site(
            "000001",
            pagamentos=pagamentos,
            acordos=acordos,
            hoje=date(2026, 7, 20),
        )

        self.assertEqual(historico["valor_em_atraso"], 100.0)
        self.assertEqual(historico["parcelas_vencidas"], 1)
        self.assertEqual(historico["valor_futuro"], 200.0)
        self.assertEqual(historico["parcelas_futuras"], 1)
        self.assertEqual(historico["valor_realizado"], 300.0)
        self.assertEqual(historico["quantidade_realizada"], 1)
        self.assertEqual(historico["valor_acordos_abertos"], 500.0)
        self.assertEqual(historico["quantidade_acordos_abertos"], 1)
        self.assertEqual(set(historico["pagamentos"]["ID SGS"]), {"V", "F", "P", "C"})

    def test_resumo_atraso_site_exclui_pagos_cancelados_e_futuros(self):
        pagamentos = pd.DataFrame([
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "V", "Status": "Pendente", "Microsiga": "123", "Data de vencimento": "2026-06-01", "Subtotal": 125.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "P", "Status": "Pago", "Microsiga": "123", "Data de vencimento": "2026-06-01", "Subtotal": 200.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "C", "Status": "Cancelado", "Microsiga": "123", "Data de vencimento": "2026-06-01", "Subtotal": 300.0},
            {**{c: "" for c in fs.PAYMENT_COLUMNS}, "ID SGS": "F", "Status": "Pendente", "Microsiga": "123", "Data de vencimento": "2026-08-01", "Subtotal": 400.0},
        ])
        resumo = fs.resumo_atraso_site(
            "000123", pagamentos=pagamentos, hoje=date(2026, 7, 20)
        )
        self.assertEqual(resumo, {
            "valor_em_atraso": 125.0,
            "parcelas_vencidas": 1,
        })

    def test_vincula_site_do_cadastro_completo_mesmo_ausente_do_snmpc(self):
        cadastro = pd.DataFrame([{
            "CÓDIGO MICROSIGA": "91187",
            "CÓDIGO AQUILES": "94733",
            "SMNPC": "AMERICA_II_BH_94733_IP",
            "NOME": "AMÉRICA II",
            "Status": "Ativo",
        }])
        pagamentos = pd.DataFrame([{
            **{c: "" for c in fs.PAYMENT_COLUMNS},
            "ID SGS": "AMERICA-II",
            "Status": "Pendente",
            "Favorecido": "ED. AMERICA II091187",
            "Data de vencimento": "2026-08-01",
            "Subtotal": 1000.0,
            "Site localizado": "Não",
        }])

        vinculados = fs.enriquecer_vinculos_financeiros(
            pagamentos,
            sites={},
            cadastro_sites=cadastro,
        )
        registro = vinculados.iloc[0]
        self.assertEqual(registro["Microsiga"], "091187")
        self.assertEqual(registro["Código Aquiles"], "94733")
        self.assertEqual(registro["Nome SNMPc"], "AMERICA_II_BH_94733_IP")
        self.assertEqual(registro["Nome Site"], "AMÉRICA II")
        self.assertEqual(registro["Site localizado"], "Sim")

        conciliacao = fs.analisar_conciliacao_financeira(
            {},
            pagamentos=pagamentos,
            acordos=pd.DataFrame(columns=fs.AGREEMENT_COLUMNS),
            cadastro_sites=cadastro,
        )
        self.assertNotIn(
            "Código Microsiga não encontrado",
            set(conciliacao.get("Tipo do problema", pd.Series(dtype=str))),
        )

    def test_normaliza_variacoes_do_codigo_microsiga(self):
        self.assertEqual(fs.normalizar_codigo_microsiga("91187"), "091187")
        self.assertEqual(fs.normalizar_codigo_microsiga("091187"), "091187")
        self.assertEqual(fs.normalizar_codigo_microsiga(91187.0), "091187")

    def test_exporta_conciliacao_excel(self):
        dados = pd.DataFrame([{
            coluna: (10.0 if coluna == "Valor" else "Teste")
            for coluna in fs.CONCILIATION_COLUMNS
        }])
        conteudo = fs.exportar_conciliacao_financeira_excel(dados)
        self.assertTrue(conteudo.startswith(b"PK"))


if __name__ == "__main__":
    unittest.main()
